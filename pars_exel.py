import openpyxl

file = openpyxl.load_workbook('Расписание 14.10-26.10 2024-2025.xlsx')

KURSES = file.sheetnames



def kurs1(num_kurs:str):
    sheet = file['1 курс']
    kurs1 = {}
    for column in range(0, sheet.max_column):
        if sheet[6][column].value != None:
            kurs1[f'{sheet[6][column].value}'] = column
    return kurs1


# sheet = file['2 курс']
# kurs2 = {}
# for column in range(0, sheet.max_column):
#     if sheet[5][column].value != None:
#         kurs2[f'{sheet[5][column].value}'] = column

# test = input('aaa: ')
# for row in range(7, sheet.max_row):
#     print(sheet[row][kurs2[test]].value)
print(KURSES)
all_kurses = input('выберите курс: ')
if KURSES[0] == all_kurses:
    print(kurs1(all_kurses))



# for kurses in file.sheetnames:
#     print(kurses)