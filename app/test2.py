import openpyxl
import re

from datetime import datetime


file = openpyxl.load_workbook('Расписание.xlsx')

text = 'На сегодня  ИСИП-28Б'

def kours(message):
    year = ''
    for i in file.sheetnames:
        sheet = file[i]
        if i == '2 курс':
            for column in range(0, sheet.max_column):
                if sheet[5][column].value != None:
                    if sheet[5][column].value == message.split('  ')[1]:
                        year = i
        elif i == '3, 4 курс':
            for column in range(0, sheet.max_column):
                if sheet[5][column].value != None:
                    if sheet[5][column].value == message.split('  ')[1]:
                        year = i
        elif i == '1 курс':
            for column in range(0, sheet.max_column):
                if sheet[6][column].value != None:
                    if sheet[6][column].value == message.split('  ')[1]:
                        year = i
    return year

def import_day(year: str)->dict:
    kurs = {}
    if year == '2 курс' or year == '3, 4 курс':
        sheet = file[year]
        for column in range(0, sheet.max_column):
            if sheet[5][column].value != None:
                kurs[f'{sheet[5][column].value}'] = column
    elif year == '1 курс':
        sheet = file[year]
        for column in range(0, sheet.max_column):
            if sheet[6][column].value != None:
                kurs[f'{sheet[6][column].value}'] = column
    return kurs

    # date = datetime.now()
    # num_day = date.weekday()
def schedule_of_the_day(message: str, year:str)->list:
    faculty = import_day(kours(message))[message.split('  ')[1]]
    sheet = file[year]
    min_num1 = [9, 26, 45, 64, 83, 102]
    max_num1 = [23, 42, 61, 80, 99, 118]
    min_num2 = [8, 25, 47, 73, 94]
    max_num2 = [23, 46, 71, 90, 105]
    schedule = []
    end_schedule = []
    time = []
    end_time = []
    class_Room = []
    end_class_Room = []

    date = datetime.now()
    num_day = date.weekday()
    num_day = 0

    if year != '1 курс':
        if num_day != 5 and num_day != 6:
            for column in range(min_num2[num_day], max_num2[num_day] - 1):
                if int(date.strftime("%U")) % 2 != 0:
                    if sheet[column][2].value == None or sheet[column][2].value == 'н/ч':
                        schedule.append(sheet[column][faculty].value)
                        time.append(
                            sheet[column][1].value if sheet[column][1].value != None else sheet[column - 1][1].value)
                        class_Room.append(
                            sheet[column][faculty + 1].value if sheet[column][faculty + 1].value != None else
                            sheet[column - 1][faculty + 1].value)
                else:
                    if sheet[column][2].value == 'ч' or sheet[column][2].value == None:
                        if any(sheet[column][faculty].coordinate in merged_range for merged_range in
                               sheet.merged_cells.ranges):
                            schedule.append(sheet[column - 1][faculty].value)
                            time.append(sheet[column][1].value if sheet[column][1].value != None else sheet[column - 1][
                                1].value)
                            class_Room.append(
                                sheet[column][faculty + 1].value if sheet[column][faculty + 1].value != None else
                                sheet[column - 1][faculty + 1].value)
                        else:
                            schedule.append(sheet[column][faculty].value)
                            time.append(sheet[column][1].value if sheet[column][1].value != None else sheet[column - 1][
                                1].value)
                            class_Room.append(
                                sheet[column][faculty + 1].value if sheet[column][faculty + 1].value != None else
                                sheet[column - 1][faculty + 1].value)
        else:
            return ['Сегодня нет занятий']
    elif year == '1 курс':
        if num_day != 6:
            for column in range(min_num1[num_day], max_num1[num_day]):
                if int(date.strftime("%U")) % 2 != 0:
                    if sheet[column][2].value == None or sheet[column][2].value == 'н/ч':
                        schedule.append(sheet[column][faculty].value)
                        time.append(
                            sheet[column][1].value if sheet[column][1].value != None else sheet[column - 1][1].value)
                        class_Room.append(
                            sheet[column][faculty + 1].value if sheet[column][faculty + 1].value != None else
                            sheet[column - 1][faculty + 1].value)
                else:
                    if sheet[column][2].value == 'ч' or sheet[column][2].value == None:
                        if any(sheet[column][faculty].coordinate in merged_range for merged_range in
                               sheet.merged_cells.ranges):
                            schedule.append(sheet[column - 1][faculty].value)
                            time.append(sheet[column][1].value if sheet[column][1].value != None else sheet[column - 1][
                                1].value)
                            class_Room.append(
                                sheet[column][faculty + 1].value if sheet[column][faculty + 1].value != None else
                                sheet[column - 1][faculty + 1].value)
                        else:
                            schedule.append(sheet[column][faculty].value)
                            time.append(sheet[column][1].value if sheet[column][1].value != None else sheet[column - 1][
                                1].value)
                            class_Room.append(
                                sheet[column][faculty + 1].value if sheet[column][faculty + 1].value != None else
                                sheet[column - 1][faculty + 1].value)
        else:
            return ['Сегодня нет занятий']
    for i in range(len(schedule)):
        if schedule[i] != None:
            end_schedule.append((re.sub(r'\s+', ' ', schedule[i])))
            end_time.append(time[i])
            end_class_Room.append(class_Room[i])
    return [end_time, end_schedule, end_class_Room]

def Tomorrows_schedule(message: str, year:str)->list:
    faculty = import_day(kours(message))[message.split('  ')[1]]
    sheet = file[year]
    min_num1 = [9, 26, 45, 64, 83, 102]
    max_num1 = [23, 42, 61, 80, 99, 118]
    min_num2 = [8, 25, 47, 73, 94]
    max_num2 = [23, 46, 71, 90, 105]
    schedule = []
    time = []
    class_Room = []
    num_lesson = []
    end_schedule = []
    end_time = []
    end_class_Room = []
    end_num_lesson = []

    date = datetime.now()
    weekday = int(date.strftime("%U"))
    num_day = date.weekday()
    num_day = 0
    if num_day != 6:
        num_day += 1
    else:
        weekday += 1
        num_day = 0

    if year != '1 курс':
        if num_day != 5 and num_day != 6:
            for column in range(min_num2[num_day], max_num2[num_day] - 1):
                if weekday % 2 != 0:
                    if sheet[column][2].value == None or sheet[column][2].value == 'н/ч':

                        schedule.append(sheet[column][faculty].value)

                        time.append(
                            sheet[column][1].value if sheet[column][1].value != None else sheet[column - 1][1].value)

                        class_Room.append(
                            sheet[column][faculty + 1].value if sheet[column][faculty + 1].value != None else
                            sheet[column - 1][faculty + 1].value)


                        num_lesson.append(
                            sheet[column][3].value if sheet[column][3].value != None else sheet[column - 1][3].value)
                else:
                    if sheet[column][2].value == 'ч' or sheet[column][2].value == None:
                        if any(sheet[column][faculty].coordinate in merged_range for merged_range in
                               sheet.merged_cells.ranges):

                            schedule.append(sheet[column - 1][faculty].value)

                            time.append(
                                sheet[column][1].value if sheet[column][1].value != None else sheet[column - 1][1].value)

                            class_Room.append(
                                sheet[column][faculty + 1].value if sheet[column][faculty + 1].value != None else
                                sheet[column - 1][faculty + 1].value)

                            num_lesson.append(
                                sheet[column][3].value if sheet[column][3].value != None else sheet[column - 1][3].value)

                        else:
                            schedule.append(sheet[column][faculty].value)

                            time.append(
                                sheet[column][1].value if sheet[column][1].value != None else sheet[column - 1][1].value)

                            class_Room.append(
                                sheet[column][faculty + 1].value if sheet[column][faculty + 1].value != None else
                                sheet[column - 1][faculty + 1].value)

                            num_lesson.append(
                                sheet[column][3].value if sheet[column][3].value != None else sheet[column - 1][3].value)

        else:
            return ['Сегодня нет занятий']
    elif year == '1 курс':
        if num_day != 6:
            for column in range(min_num1[num_day], max_num1[num_day]):
                if weekday % 2 != 0:
                    if sheet[column][2].value == None or sheet[column][2].value == 'н/ч':
                        schedule.append(sheet[column][faculty].value)
                        time.append(
                            sheet[column][1].value if sheet[column][1].value != None else sheet[column - 1][1].value)
                        class_Room.append(
                            sheet[column][faculty + 1].value if sheet[column][faculty + 1].value != None else
                            sheet[column - 1][faculty + 1].value)
                else:
                    if sheet[column][2].value == 'ч' or sheet[column][2].value == None:
                        if any(sheet[column][faculty].coordinate in merged_range for merged_range in
                               sheet.merged_cells.ranges):
                            schedule.append(sheet[column - 1][faculty].value)
                            time.append(sheet[column][1].value if sheet[column][1].value != None else sheet[column - 1][
                                1].value)
                            class_Room.append(
                                sheet[column][faculty + 1].value if sheet[column][faculty + 1].value != None else
                                sheet[column - 1][faculty + 1].value)
                        else:
                            schedule.append(sheet[column][faculty].value)
                            time.append(sheet[column][1].value if sheet[column][1].value != None else sheet[column - 1][
                                1].value)
                            class_Room.append(
                                sheet[column][faculty + 1].value if sheet[column][faculty + 1].value != None else
                                sheet[column - 1][faculty + 1].value)
        else:
            return ['Сегодня нет занятий']

    for i in range(len(schedule)):
        if schedule[i] != None:
            end_schedule.append((re.sub(r'\s+', ' ', schedule[i])))
            end_time.append(time[i])
            end_class_Room.append(class_Room[i])
            end_num_lesson.append(num_lesson[i])
    return [end_num_lesson, end_time, end_schedule, end_class_Room]

name = Tomorrows_schedule(text, kours(text))
if len(name) != 1:
    for i in range(len(name[1])):
        print(name[0][i],name[1][i], name[2][i], name[3][i])
else:
    print(name[0])